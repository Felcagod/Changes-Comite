import pandas as pd
import os
import re
import calendar
import streamlit as st
import tempfile
from Gerar_planilha_comite import padronizar_e_gerar_planilha
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl import load_workbook
from openpyxl.comments import Comment

st.title("Gerador de Planilha Comitê")
st.write("Carregue os arquivos Jira e Maximo para gerar a planilha final.")

st.title("Gerador de Planilha Comitê")

jira_file = st.file_uploader("Escolha o arquivo Jira.xlsx", type=["xlsx"])
maximo_file = st.file_uploader("Escolha o arquivo Maximo.xlsx ou Maximo.csv", type=["xlsx", "csv"])

if jira_file and maximo_file:
    if st.button("Gerar planilha"):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_output:
            output_path = tmp_output.name
        
        # Salvar arquivos temporariamente para passar o caminho
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_jira:
            tmp_jira.write(jira_file.read())
            caminho_jira = tmp_jira.name
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx" if maximo_file.type=="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" else ".csv") as tmp_maximo:
            tmp_maximo.write(maximo_file.read())
            caminho_maximo = tmp_maximo.name

        try:
            planilha_gerada = padronizar_e_gerar_planilha(caminho_jira, caminho_maximo, output_path)
            st.success("Planilha gerada com sucesso!")
            with open(planilha_gerada, "rb") as f:
                st.download_button("Baixar planilha", data=f, file_name="planilha_final.xlsx")
        except Exception as e:
            st.error(f"Erro ao gerar a planilha: {e}")


def ler_jira(caminho_jira, colunas_esperadas):
    try:
        df_jira = pd.read_excel(caminho_jira, sheet_name='Your Jira Issues', header=0)
        if all(col in df_jira.columns for col in colunas_esperadas):
            return df_jira[colunas_esperadas]
        else:
            print("⚠️ Colunas esperadas não encontradas na aba 'Your Jira Issues', ignorando Jira.")
            return None
    except Exception as e:
        print(f"⚠️ Erro ao ler Jira.xlsx: {e}. Ignorando dados do Jira.")
        return None


def ler_maximo(caminho_maximo, colunas_esperadas):
    df_maximo = None

    if caminho_maximo.lower().endswith('.xlsx'):
        try:
            df = pd.read_excel(caminho_maximo, sheet_name='Maximo')
            df = df.rename(columns={
                "change_number": "Chave",
                "summary": "Resumo",
                "status": "Status",
                "details": "Descrição",
                "owner_name": "Relator",
                "schedule_start": "Planned start date",
                "schedule_finish": "Planned end date"
            })
            if all(col in df.columns for col in colunas_esperadas):
                df_maximo = df[colunas_esperadas]
            else:
                print("⚠️ Colunas esperadas não encontradas na aba 'Maximo' do Excel.")
        except Exception as e:
            print(f"⚠️ Erro ao ler Maximo.xlsx: {e}.")
    elif caminho_maximo.lower().endswith('.csv'):
        try:
            df = pd.read_csv(caminho_maximo, encoding="utf-8", sep=",")
            df = df.rename(columns={
                "change_number": "Chave",
                "summary": "Resumo",
                "status": "Status",
                "details": "Descrição",
                "owner_name": "Relator",
                "schedule_start": "Planned start date",
                "schedule_finish": "Planned end date"
            })
            if all(col in df.columns for col in colunas_esperadas):
                df_maximo = df[colunas_esperadas]
            else:
                print("⚠️ Colunas esperadas não encontradas no CSV Maximo.csv.")
        except Exception as e:
            print(f"⚠️ Erro ao ler Maximo.csv: {e}")

    if df_maximo is None:
        raise FileNotFoundError("❌ Não foi possível ler dados válidos do Maximo.")

    df_maximo["Planned start date"] = pd.to_datetime(df_maximo["Planned start date"], errors="coerce", dayfirst=True)
    df_maximo["Planned end date"] = pd.to_datetime(df_maximo["Planned end date"], errors="coerce", dayfirst=True)
    df_maximo = df_maximo[df_maximo["Chave"] != "String"]
    df_maximo = df_maximo[df_maximo["Status"] == "AUTH"]

    return df_maximo


def aplicar_formatacao_excel(caminho_arquivo, abas):
    wb = load_workbook(caminho_arquivo)

    fonte_normal = Font(name="Montserrat", size=11)
    fonte_titulo = Font(name="Montserrat", size=12, bold=True)
    thin_side = Side(border_style="thin", color="C4C7C5")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    cores = {
        "Jira": {
            "header": "8989EB",
            "even": "C4C3F7",
            "odd":  "E8E7FC",
        },
        "Maximo": {
            "header": "8BC34A",
            "even": "FFFFFF",
            "odd":  "EEF7E3",
        },
        "Participantes": {
            "header": "4B7BEC",
            "even": "E6F0FF",
            "odd":  "FFFFFF",
        },
        "Verificação": {
            "header": "FF5722",
            "even": "FFCCBC",
            "odd": "FFEDE6",
        }
    }

    comentarios_participantes = {
        "A2": "Igor Campos , Reginaldo Tadashi , Newton Albuquerque , Adilson Bassani",
        "A3": "Clodoaldo Dias , Wesley Magalhães",
        "A4": "Roberto , Wagner",
        "A5": "Sergio Massao , Wilton Carvalho",
        "A6": "Ricardo Witsmiszyn",
        "A7": "Enrique Dias , Ricardo Witsmiszyn",
        "A8": "Thais Yuta , Mauricio Souza , Guilherme Perdroso",
        "A9": "Thiago Pezzini , Samuel Silva"
    }

    for aba in abas:
        if aba not in wb.sheetnames:
            continue

        ws = wb[aba]
        cfg = cores.get(aba, {})
        fill_header = PatternFill(start_color=cfg.get("header", "CCCCCC"), fill_type="solid")
        fill_even = PatternFill(start_color=cfg.get("even", "FFFFFF"), fill_type="solid")
        fill_odd = PatternFill(start_color=cfg.get("odd", "FFFFFF"), fill_type="solid")

        if aba in ["Jira", "Maximo"]:
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row,
                                                min_col=1, max_col=ws.max_column), 1):
                for j, cell in enumerate(row, 1):
                    cell.border = thin_border

                    if i == 1:
                        cell.fill = fill_header
                        cell.font = fonte_titulo
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    else:
                        cell.font = fonte_normal
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                        if i % 2 == 0:
                            cell.fill = fill_even
                        else:
                            cell.fill = fill_odd

                        if cell.column_letter in ['F', 'G']:
                            cell.number_format = "DD/MM/YYYY HH:mm"

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                if column == 'A':
                    adjusted_width = max(max_length + 5, 20)
                else:
                    adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column].width = adjusted_width

            ws.sheet_view.showGridLines = True

        elif aba == "Participantes":
            # Limpar aba para evitar restos antigos
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
                    cell.fill = PatternFill(fill_type=None)
                    cell.border = Border()
                    cell.font = Font()
                    cell.alignment = Alignment()

            participantes_obrigatorios = [
                ["Cargo", "Responsável"],
                ["Arquitetura:", ""],
                ["Field:", ""],
                ["Governança:", ""],
                ["Infraestrutura:", ""],
                ["N1:", ""],
                ["Operação:", ""],
                ["Segurança:", ""],
                ["Telecom:", ""],
            ]

            participantes = [
                ["Participantes", "Lista"],
            ]

            for i, row_data in enumerate(participantes_obrigatorios):
                for j, value in enumerate(row_data):
                    cell = ws.cell(row=i + 1, column=j + 1, value=value)
                    if i == 0:
                        cell.fill = fill_header
                        cell.font = fonte_titulo
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    else:
                        cell.fill = fill_odd if i % 2 else fill_even
                        cell.font = fonte_normal
                        if j == 0:
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                    cell.border = thin_border

            for coord, texto in comentarios_participantes.items():
                cell = ws[coord]
                cell.comment = Comment(texto, "GPT")

            start_row_segunda_tabela = len(participantes_obrigatorios) + 4

            for i, row_data in enumerate(participantes):
                for j, value in enumerate(row_data):
                    cell = ws.cell(start_row_segunda_tabela + i, column=j + 1, value=value)
                    if i == 0:
                        cell.fill = fill_header
                        cell.font = fonte_titulo
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    else:
                        cell.fill = fill_odd if i % 2 else fill_even
                        cell.font = fonte_normal
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.border = thin_border

            max_colunas = max(len(participantes_obrigatorios[0]), len(participantes[0]))
            for col_idx in range(1, max_colunas + 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for row_idx in range(1, start_row_segunda_tabela + len(participantes) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max(max_length + 5, 15), 50)
                ws.column_dimensions[col_letter].width = adjusted_width

            ws.sheet_view.showGridLines = True

        elif aba == "Verificação":
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row,
                                                min_col=1, max_col=ws.max_column), 1):
                for cell in row:
                    cell.border = thin_border
                    if i == 1:
                        cell.fill = fill_header
                        cell.font = Font(name="Montserrat", size=13, bold=True, color="FFFFFF")
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        cell.fill = fill_even if i % 2 == 0 else fill_odd
                        cell.font = fonte_normal
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                    if cell.column_letter in ['F', 'G']:
                        cell.number_format = "DD/MM/YYYY HH:mm"

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                if column == 'A':
                    adjusted_width = max(max_length + 8, 30)
                else:
                    adjusted_width = min(max(max_length + 3, 15), 60)
                ws.column_dimensions[column].width = adjusted_width

            ws.sheet_view.showGridLines = True

    wb.save(caminho_arquivo)
    print(f"✅ Formatação aplicada nas abas: {', '.join(abas)}")


def is_data_relevante(data):
    if pd.isna(data):
        return False
    ano = data.year
    mes = data.month
    dia = data.day

    ultimo_dia = calendar.monthrange(ano, mes)[1]

    if dia == ultimo_dia:
        return True

    if mes == 12:
        proximo_mes = 1
        proximo_ano = ano + 1
    else:
        proximo_mes = mes + 1
        proximo_ano = ano

    if dia == 1 and mes == proximo_mes and ano == proximo_ano:
        return True

    return False


def padronizar_e_gerar_planilha(caminho_jira, caminho_maximo, nome_saida='planilha_final.xlsx'):
    colunas_finais = [
        "Chave", "Resumo", "Status", "Descrição", "Relator",
        "Planned start date", "Planned end date"
    ]

    df_jira = ler_jira(caminho_jira, colunas_finais)
    df_maximo = ler_maximo(caminho_maximo, colunas_finais)

    dfs = [df for df in [df_jira, df_maximo] if df is not None]
    df_completo = pd.concat(dfs, ignore_index=True)

    sistemas_criticos = [
        "ARS\\NCR",
        "Athena",
        "Concentrador Fiscal",
        "Concsitef",
        "CTF",
        "Gescom",
        "Gold",
        "Guepardo",
        "MasterSaf",
        "Pegasus Descontos Comerciais",
        "SAD Contábil",
        "SAP",
        "SCE",
        "Sitef",
        "Storex",
        "TPLinux",
        "XRT"
    ]

    regex_sistemas = re.compile("|".join([re.escape(s) for s in sistemas_criticos]), re.IGNORECASE)

    filtro_sistemas = df_completo.apply(
        lambda row: bool(regex_sistemas.search(str(row["Resumo"]))) or bool(regex_sistemas.search(str(row["Descrição"]))),
        axis=1
    )

    def filtro_data(row):
        return is_data_relevante(row["Planned start date"]) or is_data_relevante(row["Planned end date"])

    df_filtrado = df_completo[filtro_sistemas].copy()
    df_filtrado = df_filtrado[df_filtrado.apply(filtro_data, axis=1)]

    abas_criadas = []

    with pd.ExcelWriter(nome_saida, engine='openpyxl') as writer:
        if df_jira is not None:
            df_jira.to_excel(writer, sheet_name='Jira', index=False)
            abas_criadas.append('Jira')
        else:
            print("⚠️ Dados Jira não foram encontrados ou estão inválidos. Aba 'Jira' não será criada.")

        df_maximo.to_excel(writer, sheet_name='Maximo', index=False)
        abas_criadas.append('Maximo')

        df_dummy = pd.DataFrame([[""]])
        df_dummy.to_excel(writer, sheet_name='Participantes', index=False, header=False)
        abas_criadas.append('Participantes')

        if not df_filtrado.empty:
            df_filtrado.to_excel(writer, sheet_name='Verificação', index=False)
            abas_criadas.append('Verificação')
        else:
            print("ℹ️ Nenhuma change em conflito com fechamento contábil. Aba 'Verificação' não será criada.")

    aplicar_formatacao_excel(nome_saida, abas_criadas)

    return nome_saida

